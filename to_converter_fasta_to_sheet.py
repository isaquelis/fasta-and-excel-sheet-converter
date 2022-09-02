from Bio import SeqIO
from openpyxl import Workbook
import re

fasta_file = "human_sequence.fasta"
sequence = [i for i in SeqIO.parse(fasta_file, 'fasta')]  # Using List Comprehension to create a SeqRecord List.

seq_wb = Workbook()  # Creating a Excel workbook.
seq_sheet = seq_wb.active  # Activating workbook as a sheet object.

sheet_header = ['Name/ID', 'Gene Name', 'Organism', 'Length', 'Sequence']
for i in range(1, 6):
    seq_sheet.cell(row=1, column=i).value = sheet_header[i - 1]  # Filling the first sheet line with the header.

for i in range(2, len(sequence) + 2):
    # Splitting FASTA description to make different columns.
    sequence_slices = re.split('[|]*[|]\s', sequence[i - 2].description)

    # Conditional that verify if the description slice is empty.
    if len(sequence_slices[1][6:len(sequence_slices[1]) - 1]) > 0:
        gene_name = sequence_slices[1][6:len(sequence_slices[1]) - 1]
    else:
        gene_name = 'Data does not exist'

    # Removes the FASTA tags and insert the sequence ID to its respective position in the column 1.    
    seq_sheet.cell(row=i, column=1).value = re.sub('head[|]', "", sequence[i - 2].id[:len(sequence[i - 2].id) - 1])

    # Insert the gene name to its respective position in the column 2.
    seq_sheet.cell(row=i, column=2).value = gene_name

    # Insert the organism name to its respective position in the column 3.
    seq_sheet.cell(row=i, column=3).value = sequence_slices[2][10:len(sequence_slices[2]) - 1]

    # Insert the sequence length to its respective position in the column 4.
    seq_sheet.cell(row=i, column=4).value = sequence_slices[3][8:len(sequence_slices[3]) - 1]

    # Insert the sequence to its respective position in the column 5.
    seq_sheet.cell(row=i, column=5).value = str(sequence[i - 2].seq)

xlsx_name = input("Define Excel-file name: ")
seq_wb.save(xlsx_name + ".xlsx")  # Saving workbook with the given input name.
