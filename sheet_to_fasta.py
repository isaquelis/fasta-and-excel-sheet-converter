import openpyxl as xl
from Bio import SeqIO
from Bio.Seq import Seq
from Bio.SeqRecord import SeqRecord

seqFile = "human_sequence.xlsx"

wb_seq = xl.load_workbook(seqFile) # To load a Excel sheet.

sheet_seq = wb_seq.active # To activate the project as a sheet object.

sequencesList = []
for i in range(2, sheet_seq.max_row + 1):
    sequencesList.append(SeqRecord(
        Seq(sheet_seq.cell(row = i, column = 8).value),
        id = ('head|' + sheet_seq.cell(row = i, column = 1).value + '|'),
        name = ('[gene='+ sheet_seq.cell(row = i, column = 5).value + ']'),
        description = ('[organism=' + sheet_seq.cell(row = i, column = 6).value + '] ' + '[length=' + sheet_seq.cell(row = i, column = 7).value + ']'),
    ))

fastaFileName = input("Define FASTA-file name: ")
SeqIO.write(sequencesList, fastaFileName + ".fasta", "fasta") # To write FASTA file with a seqRecord objects list.