from Bio import SeqIO
from openpyxl import Workbook as wb

FastaFile = "human_sequence.fasta"

sequence = [i for i in SeqIO.parse(FastaFile, 'fasta')]

seqWb = wb()
seqSheet = seqWb.active

sheetHead = ['Name-ID', 'Description', 'Sequence']
for i in range(1, 4):
    seqSheet.cell(row = 1, column = i).value = sheetHead[i-1]

for i in range(2, len(sequence)+1):
    seqSheet.cell(row = i, column = 1).value = sequence[i-2].id
    seqSheet.cell(row = i, column = 2).value = sequence[i-2].description[len(sequence[0].id)+1:]
    seqSheet.cell(row = i, column = 3).value = str(sequence[i-2].seq)

xlsxOutputName = input("Define Excel-file name: ")
seqWb.save(xlsxOutputName + ".xlsx")