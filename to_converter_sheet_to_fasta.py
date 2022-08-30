import openpyxl as xl
from Bio import SeqIO
from Bio.Seq import Seq
from Bio.SeqRecord import SeqRecord

seq_file = "human_sequence.xlsx"
wb_seq = xl.load_workbook(seq_file) # To load a Excel sheet.

sheet_seq = wb_seq.active # To activate the project as a sheet object.

sequences_list = []
for i in range(2, sheet_seq.max_row + 1):
   
   # Line 21: Selecting sequence of a SeqRecord object.
   # Line 22: Adding header tittle and sequence ID.
   # Line 23: Adding gene name to its gap into description.
   # Line 24: Adding organism name to its gap into description.
   # Line 25: Adding sequence length to its gap into description.
   
    sequences_list.append(SeqRecord(
        Seq(sheet_seq.cell(row = i, column = 8).value),
        id = ('head|' + sheet_seq.cell(row = i, column = 1).value + '|'),
        description = ('[gene=' + sheet_seq.cell(row = i, column = 5).value + ']| ' +
                       '[organism=' + sheet_seq.cell(row = i, column = 6).value + ']| ' +
                       '[length=' + sheet_seq.cell(row = i, column = 7).value + ']')
    ))

fasta_name = input("Define FASTA-file name: ")
SeqIO.write(sequences_list, fasta_name + ".fasta", "fasta") # To write FASTA file with a seqRecord objects list.
